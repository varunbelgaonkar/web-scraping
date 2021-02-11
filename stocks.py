from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import os
import time

url = 'https://www.moneycontrol.com/'

companies = []
cmp = []

#creating a spreadsheet file
worksheet = "stocks.xlsx"
def create_worksheet():
	if not os.path.exists(worksheet):
		workbook = Workbook()
		sheet = workbook.active
		sheet["A1"] = "COMPANY NAME"
		sheet["B1"] = "CMP"
		sheet["C1"] = "AVG"
		sheet["D1"] = "Quantity"
		sheet["E1"] = "Investment"
		sheet["F1"] = "P/L"
		sheet["G1"] = "%P/L"
		workbook.save(worksheet)
	else:
		print("File already exists!!")

#reading the spreadsheet 
def read_worksheet():
	workbook = load_workbook(worksheet)
	sheet = workbook.active
	
	for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1,values_only=True):
		companies.append(row[0])
	return companies


#automating the browser
def current_price():
	
	companies = read_worksheet()
	driver = webdriver.Chrome(executable_path=r"C:\Users\varun\OneDrive\Documents\python projects\chromedriver.exe")
	driver.get(url)
	page = driver.page_source
	for company in companies:
		driver.find_element_by_name('search_str').send_keys(company)
		driver.find_element_by_name('search_str').send_keys(Keys.ENTER)
		price = driver.find_element_by_id('nsecp').get_attribute("data-numberanimate-value")
		#print(price)
		cmp.append((company,price))
		time.sleep(3)
	return cmp

#current_price()

#writing data to worksheet
def write_worksheet():
	s_row = 2
	workbook = load_workbook(worksheet)
	sheet = workbook.active
	cmp = current_price()
	for c,p in cmp:
		sheet.cell(row=s_row,column=2,value=p)
		s_row += 1
	workbook.save(worksheet)


write_worksheet()





		




